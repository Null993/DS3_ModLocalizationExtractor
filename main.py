#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
menu_tool_batch.py

Two main features:
 - Extraction: split original FromSoftware JSON into header (0_header.json) + parts
 - Merge: read header + parts and restore original JSON structure

Supported part export formats:
 - 极简默认 (minimal): part_x.json -> {"0": "text0", "1": "text1", ...}
 - MTool JSON (mtool): part_x.json -> {"原文1": "初始译文1", ...}  (initial translation = original)
 - Translator++ XLSX (translatorpp): part_x.xlsx with header columns including OriginalText and Initial (Initial left empty)
"""

import json
import os
import math
import threading
import datetime
from difflib import SequenceMatcher
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# optional dependency
try:
    import openpyxl
except Exception:
    openpyxl = None

# ---------------------------
# Utilities
# ---------------------------
def safe_json_load(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def safe_json_dump(obj, path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)

def estimate_tokens(text: str) -> int:
    if not text:
        return 1
    return max(1, math.ceil(len(text) / 4.0))

def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()

# ---------------------------
# Build structure / EntryIDs
# ---------------------------
def build_structure_and_entryids(json_data):
    """
    Extracts:
      - top_name (json_data["Name"])
      - structure: list of dicts each containing:
            WrapperName, WrapperID, FmgName, EntryIndexes, FmgVersion, FmgBigEndian, FmgUnicode, FmgCompression
      - entry_ids: list of all Entry IDs in order
    """
    top_name = json_data.get("Name", "")
    structure = []
    entry_ids = []

    for wrapper in json_data.get("FmgWrappers", []):
        wname = wrapper.get("Name", "")
        wid = wrapper.get("ID", 0)
        fmg = wrapper.get("Fmg", {}) or {}
        fname = fmg.get("Name", "")

        # FMG meta
        fmg_version = fmg.get("Version")
        fmg_bigendian = fmg.get("BigEndian")
        fmg_unicode = fmg.get("Unicode")
        fmg_compression = fmg.get("Compression")

        indexes = []
        for entry in fmg.get("Entries", []):
            idx = len(entry_ids)
            indexes.append(idx)
            entry_ids.append(entry.get("ID"))

        structure.append({
            "WrapperName": wname,
            "WrapperID": wid,
            "FmgName": fname,
            "EntryIndexes": indexes,
            "FmgVersion": fmg_version,
            "FmgBigEndian": fmg_bigendian,
            "FmgUnicode": fmg_unicode,
            "FmgCompression": fmg_compression
        })

    return top_name, structure, entry_ids

# ---------------------------
# Write header + parts
# ---------------------------
# ---------- 替换：write_header_and_parts ----------
def write_header_and_parts(source_file, top_name, structure, entry_ids, texts, split, max_per_file, extract_format):
    """
    Write 0_header.json and parts according to extract_format.
    extract_format values: "极简默认", "mtool", "translatorpp"
    - 极简默认 -> JSON array: ["t0","t1","",...]
    - translatorpp -> XLSX with columns [OriginalText, Initial] (Initial empty); fallback to JSON array if openpyxl missing
    """
    base_dir = os.path.dirname(source_file)
    base_name = os.path.splitext(os.path.basename(source_file))[0]
    out_dir = os.path.join(base_dir, base_name + "_extracted")
    os.makedirs(out_dir, exist_ok=True)

    # build chunks metadata
    if not split:
        chunks = [{"start": 0, "count": len(texts)}]
    else:
        chunks = []
        start = 0
        for i in range(0, len(texts), max_per_file):
            part = texts[i:i + max_per_file]
            chunks.append({"start": start, "count": len(part)})
            start += len(part)

    header = {
        "Meta": {
            "SourceTopName": top_name,
            "Version": 1,
            "TotalEntries": len(texts),
            "ChunkCount": len(chunks),
            "Chunks": chunks,
            "ExtractFormat": extract_format  # record chosen format
        },
        "Structure": structure,
        "EntryIDs": entry_ids
    }

    header_path = os.path.join(out_dir, "0_header.json")
    safe_json_dump(header, header_path)

    # write parts according to new rules
    for idx, ch in enumerate(chunks, start=1):
        start = ch["start"]
        count = ch["count"]
        slice_texts = texts[start:start + count]

        if extract_format == "极简默认":
            # JSON array: ["t0","t1","",...]
            part_path = os.path.join(out_dir, f"part_{idx}.json")
            safe_json_dump(slice_texts, part_path)

        elif extract_format == "translatorpp":
            # Write XLSX with only two columns: OriginalText, Initial (Initial empty)
            if openpyxl is None:
                # fallback to JSON array to preserve order
                part_path = os.path.join(out_dir, f"part_{idx}.json")
                safe_json_dump(slice_texts, part_path)
            else:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Entries"
                # Only write two columns as requested
                ws.append(["OriginalText", "Initial"])
                for txt in slice_texts:
                    ws.append([txt if txt is not None else "", ""])
                part_path = os.path.join(out_dir, f"part_{idx}.xlsx")
                wb.save(part_path)
        else:
            # unknown -> fallback to minimal JSON array
            part_path = os.path.join(out_dir, f"part_{idx}.json")
            safe_json_dump(slice_texts, part_path)

    return len(chunks), header_path, out_dir

# ---------------------------
# Merge folder: read header + parts -> merged entries
# ---------------------------
# ---------- 替换：merge_folder ----------
def merge_folder(folder_path):
    header_path = os.path.join(folder_path, "0_header.json")
    if not os.path.exists(header_path):
        raise FileNotFoundError("0_header.json 未找到。")

    header = safe_json_load(header_path)
    meta = header.get("Meta", {})
    top_name = meta.get("SourceTopName", "")
    total = int(meta.get("TotalEntries", 0))
    chunks = meta.get("Chunks", [])
    extract_format = meta.get("ExtractFormat", "极简默认")

    structure = header.get("Structure", [])
    entry_ids = header.get("EntryIDs", [])

    # prepare blank texts
    texts = [""] * total

    # normalize key for decision
    fmt = extract_format
    if fmt == "极简默认" or fmt == "minimal":
        key_fmt = "minimal"
    elif isinstance(fmt, str) and fmt.lower().startswith("translator"):
        key_fmt = "translatorpp"
    else:
        key_fmt = fmt

    for idx, ch in enumerate(chunks, start=1):
        start = ch.get("start", 0)
        count = ch.get("count", 0)
        json_path = os.path.join(folder_path, f"part_{idx}.json")
        xlsx_path = os.path.join(folder_path, f"part_{idx}.xlsx")

        if key_fmt == "minimal":
            # Accept both new JSON array and old dict style for backwards compatibility
            if not os.path.exists(json_path):
                raise FileNotFoundError(f"缺失分片文件：{os.path.basename(json_path)}")
            pdata = safe_json_load(json_path)
            if isinstance(pdata, list):
                # new minimal array format
                for k in range(count):
                    texts[start + k] = pdata[k] if k < len(pdata) else ""
            elif isinstance(pdata, dict):
                # old dict with numeric keys
                for k in range(count):
                    texts[start + k] = pdata.get(str(k), "")
            else:
                for k in range(count):
                    texts[start + k] = ""

        elif key_fmt == "translatorpp":
            # Prefer XLSX; fallback to JSON array.
            if os.path.exists(xlsx_path) and openpyxl is not None:
                wb = openpyxl.load_workbook(xlsx_path, read_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    for k in range(count):
                        texts[start + k] = ""
                else:
                    header_row = [ (c or "").strip() for c in rows[0] ]
                    # find OriginalText and Initial indices (case-insensitive tolerant)
                    orig_idx = None
                    init_idx = None
                    for i, col in enumerate(header_row):
                        low = col.lower()
                        if low.replace(" ", "") in ("originaltext", "originaltext"):
                            orig_idx = i
                        if low.replace(" ", "") in ("initial", "initialtranslation", "initialtranslation"):
                            init_idx = i
                    if orig_idx is None:
                        orig_idx = 0
                    if init_idx is None:
                        init_idx = 1 if len(header_row) > 1 else None
                    data_rows = rows[1:]
                    for k in range(count):
                        if k < len(data_rows):
                            row = data_rows[k]
                            orig = row[orig_idx] if orig_idx < len(row) and row[orig_idx] is not None else ""
                            ini = ""
                            if init_idx is not None and init_idx < len(row) and row[init_idx] is not None:
                                ini = row[init_idx]
                            texts[start + k] = str(ini) if (ini is not None and str(ini).strip() != "") else str(orig)
                        else:
                            texts[start + k] = ""
            else:
                # fallback to JSON array or other structures
                if not os.path.exists(json_path):
                    raise FileNotFoundError(f"缺失分片文件：{os.path.basename(json_path)}")
                pdata = safe_json_load(json_path)
                if isinstance(pdata, list):
                    # assume it's array of original texts
                    for k in range(count):
                        texts[start + k] = pdata[k] if k < len(pdata) else ""
                elif isinstance(pdata, dict) and "entries" in pdata:
                    arr = pdata.get("entries", [])
                    for k in range(count):
                        texts[start + k] = arr[k].get("text", "") if k < len(arr) else ""
                elif isinstance(pdata, dict):
                    # numeric keys fallback
                    if all(str(x).isdigit() for x in pdata.keys()):
                        for k in range(count):
                            texts[start + k] = pdata.get(str(k), "")
                    else:
                        # dict of original->initial values
                        vals = list(pdata.values())
                        for k in range(count):
                            texts[start + k] = vals[k] if k < len(vals) else ""
                elif isinstance(pdata, list):
                    for k in range(count):
                        if k < len(pdata):
                            item = pdata[k]
                            if isinstance(item, dict):
                                texts[start + k] = item.get("text") or item.get("Text") or ""
                            else:
                                texts[start + k] = str(item)
                        else:
                            texts[start + k] = ""
                else:
                    for k in range(count):
                        texts[start + k] = ""

        else:
            # best-effort for unknown
            json_path = os.path.join(folder_path, f"part_{idx}.json")
            if not os.path.exists(json_path):
                raise FileNotFoundError(f"缺失分片文件：{os.path.basename(json_path)}")
            pdata = safe_json_load(json_path)
            if isinstance(pdata, list):
                for k in range(count):
                    texts[start + k] = pdata[k] if k < len(pdata) else ""
            elif isinstance(pdata, dict):
                if all(k.isdigit() for k in pdata.keys()):
                    for k in range(count):
                        texts[start + k] = pdata.get(str(k), "")
                else:
                    vals = list(pdata.values())
                    for k in range(count):
                        texts[start + k] = vals[k] if k < len(vals) else ""
            else:
                for k in range(count):
                    texts[start + k] = ""

    # build merged entries
    merged_entries = []
    for i in range(len(entry_ids)):
        eid = entry_ids[i] if i < len(entry_ids) else None
        merged_entries.append({"ID": eid, "Text": texts[i] if i < len(texts) else ""})

    return top_name, structure, merged_entries

# ---------------------------
# Restore original JSON structure
# ---------------------------
def restore_original_from_parts(top_name, structure, merged_entries):
    lookup = {i: e["Text"] for i, e in enumerate(merged_entries)}
    entry_ids = [e["ID"] for e in merged_entries]

    wrappers = []
    for block in structure:
        entry_list = []
        for idx in block.get("EntryIndexes", []):
            eid = entry_ids[idx] if 0 <= idx < len(entry_ids) else None
            entry_list.append({"ID": eid, "Text": lookup.get(idx, "")})

        fmg_obj = {
            "Name": block.get("FmgName", ""),
            "Entries": entry_list,
            "Version": block.get("FmgVersion"),
            "BigEndian": block.get("FmgBigEndian"),
            "Unicode": block.get("FmgUnicode"),
            "Compression": block.get("FmgCompression")
        }
        wrappers.append({
            "Name": block.get("WrapperName", ""),
            "ID": block.get("WrapperID", 0),
            "Fmg": fmg_obj
        })

    return {"Name": top_name, "FmgWrappers": wrappers}

# ---------------------------
# GUI Application
# ---------------------------
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("FromSoft JSON Tool (Extraction & Merge)")
        self.root.geometry("940x620")

        # Variables
        self.extract_file_var = tk.StringVar()
        self.extract_split_var = tk.IntVar(value=0)
        self.extract_max_var = tk.StringVar(value="250")
        self.extract_format_var = tk.StringVar(value="极简默认")

        self.merge_folder_var = tk.StringVar()
        self.merge_savevar = tk.StringVar()

        # Lock for IO operations
        self._io_lock = threading.Lock()

        self.build_ui()

    def build_ui(self):
        nb = ttk.Notebook(self.root)
        nb.pack(fill="both", expand=True, padx=8, pady=8)

        tab_extract = tk.Frame(nb)
        tab_merge = tk.Frame(nb)

        nb.add(tab_extract, text="提取 (Extraction)")
        nb.add(tab_merge, text="合并 (Merge)")

        self.build_extract_tab(tab_extract)
        self.build_merge_tab(tab_merge)

    # ---------- Extract Tab ----------
    def build_extract_tab(self, tab):
        padx = 12
        tk.Label(tab, text="选择原始 JSON 文件：").pack(anchor="w", padx=padx, pady=(10,0))
        frm = tk.Frame(tab); frm.pack(fill="x", padx=padx, pady=(4,0))
        tk.Entry(frm, textvariable=self.extract_file_var, width=86).pack(side="left", fill="x", expand=True)
        tk.Button(frm, text="浏览", command=self.on_browse_extract).pack(side="left", padx=6)

        tk.Checkbutton(tab, text="拆分多个分片", variable=self.extract_split_var).pack(anchor="w", padx=padx, pady=(8,0))

        subfrm = tk.Frame(tab); subfrm.pack(anchor="w", padx=padx, pady=(6,0))
        tk.Label(subfrm, text="每个分片最大条数：").pack(side="left")
        tk.Entry(subfrm, textvariable=self.extract_max_var, width=10).pack(side="left", padx=6)

        fmtfrm = tk.Frame(tab); fmtfrm.pack(anchor="w", padx=padx, pady=(8,0))
        tk.Label(fmtfrm, text="导出格式：").pack(side="left")
        cb = ttk.Combobox(fmtfrm, textvariable=self.extract_format_var, state="readonly", width=28)
        cb['values'] = ["极简默认", "translatorpp"]
        cb.pack(side="left", padx=6)
        tk.Label(fmtfrm, text="(translator++ -> XLSX)").pack(side="left", padx=8)

        tk.Button(tab, text="开始提取", command=self.on_start_extract, height=2).pack(pady=12)
        tk.Label(tab, text="提示：提取后会在源文件同级生成 <文件名>_extracted 目录，头文件为 0_header.json").pack(anchor="w", padx=padx, pady=(6,0))

    def on_browse_extract(self):
        p = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
        if p:
            self.extract_file_var.set(p)

    def on_start_extract(self):
        p = self.extract_file_var.get()
        if not p:
            messagebox.showerror("错误", "请选择原始 JSON 文件")
            return
        if not os.path.exists(p):
            messagebox.showerror("错误", "所选文件不存在")
            return

        try:
            data = safe_json_load(p)
        except Exception as e:
            messagebox.showerror("错误", f"读取 JSON 失败：{e}")
            return

        top_name, structure, entry_ids = build_structure_and_entryids(data)

        texts = []
        for wrapper in data.get("FmgWrappers", []):
            fmg = wrapper.get("Fmg", {}) or {}
            for entry in fmg.get("Entries", []):
                texts.append(entry.get("Text", ""))

        split_flag = bool(self.extract_split_var.get())
        try:
            maxn = int(self.extract_max_var.get()) if split_flag else 0
        except:
            messagebox.showerror("错误", "分片大小必须为整数")
            return

        fmt_choice = self.extract_format_var.get() or "极简默认"

        # if translatorpp chosen but openpyxl missing, ask to continue with JSON fallback
        if fmt_choice == "translatorpp" and openpyxl is None:
            ok = messagebox.askyesno("openpyxl 未安装", "你选择了 Translator++ (XLSX) 导出，但当前未安装 openpyxl。程序将退回写入 JSON 分片作为替代。是否继续？")
            if not ok:
                return

        def worker():
            with self._io_lock:
                try:
                    count, header_path, out_dir = write_header_and_parts(
                        p, top_name, structure, entry_ids, texts, split_flag, maxn, fmt_choice
                    )
                    # set default merge folder
                    self.merge_folder_var.set(out_dir)
                    messagebox.showinfo("完成", f"提取完成：已生成头文件与 {count} 个分片\n目录：{out_dir}")
                except Exception as e:
                    messagebox.showerror("写入失败", f"{e}")

        threading.Thread(target=worker, daemon=True).start()

    # ---------- Merge Tab ----------
    def build_merge_tab(self, tab):
        padx = 12
        tk.Label(tab, text="选择分片文件夹（包含 0_header.json 与 part_*.json/xlsx）：").pack(anchor="w", padx=padx, pady=(10,0))
        frm = tk.Frame(tab); frm.pack(fill="x", padx=padx, pady=(4,0))
        tk.Entry(frm, textvariable=self.merge_folder_var, width=86).pack(side="left", fill="x", expand=True)
        tk.Button(frm, text="浏览", command=self.on_browse_merge_folder).pack(side="left", padx=6)

        tk.Label(tab, text="输出文件（可点击选择）：").pack(anchor="w", padx=padx, pady=(8,0))
        frm2 = tk.Frame(tab); frm2.pack(fill="x", padx=padx, pady=(4,0))
        tk.Entry(frm2, textvariable=self.merge_savevar, width=86).pack(side="left", fill="x", expand=True)
        tk.Button(frm2, text="选择输出", command=self.on_choose_merge_save).pack(side="left", padx=6)

        tk.Button(tab, text="开始合并并恢复", command=self.on_start_merge, height=2).pack(pady=12)
        tk.Label(tab, text="提示：程序会读取 0_header.json 中的 ExtractFormat 自动决定如何解析分片。").pack(anchor="w", padx=padx, pady=(6,0))

    def on_browse_merge_folder(self):
        p = filedialog.askdirectory()
        if p:
            self.merge_folder_var.set(p)
            foldername = os.path.basename(os.path.normpath(p))
            default_out = os.path.join(os.path.dirname(p), foldername + "_merged.json")
            self.merge_savevar.set(default_out)

    def on_choose_merge_save(self):
        initial = self.merge_savevar.get() or os.getcwd()
        init_dir = os.path.dirname(initial) if os.path.dirname(initial) else None
        fname = os.path.basename(initial) if os.path.basename(initial) else None
        p = filedialog.asksaveasfilename(defaultextension=".json", initialfile=fname, initialdir=init_dir, filetypes=[("JSON files","*.json")])
        if p:
            self.merge_savevar.set(p)

    def on_start_merge(self):
        folder = self.merge_folder_var.get()
        savep = self.merge_savevar.get()
        if not folder:
            messagebox.showerror("错误", "请选择分片文件夹")
            return
        if not os.path.isdir(folder):
            messagebox.showerror("错误", "所选路径不是文件夹")
            return
        if not savep:
            foldername = os.path.basename(os.path.normpath(folder))
            savep = os.path.join(os.path.dirname(folder), foldername + "_merged.json")
            self.merge_savevar.set(savep)

        def worker():
            with self._io_lock:
                try:
                    top_name, structure, merged_entries = merge_folder(folder)
                    restored = restore_original_from_parts(top_name, structure, merged_entries)
                    safe_json_dump(restored, savep)
                    messagebox.showinfo("完成", f"合并并恢复完成：\n{savep}")
                except Exception as e:
                    messagebox.showerror("合并失败", f"{e}")

        threading.Thread(target=worker, daemon=True).start()

# ---------------------------
# Main
# ---------------------------
def main():
    root = tk.Tk()
    app = App(root)
    root.mainloop()

if __name__ == "__main__":
    main()
